"""Unit tests for Excel I/O utilities."""

# Module responsibilities:
# - Validate happy path writing with automatic pagination.
# - Assert defensive behaviour when required columns are missing.

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from autoflow_io.excel_writer import write_fixed
from autoflow_io.mapping import FixedMapping, MappingError
from autoflow_io.schema import TargetSchema


def _build_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    for _ in range(9):
        ws.append([None, None, None])
    headers = ["项目名称", "数量", "金额(USD)"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=9, column=idx, value=header)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_mapping(path: Path, max_rows: int = 2) -> FixedMapping:
    payload = (
        "sheet: Invoice\n"
        "start_row: 10\n"
        "header_row: 9\n"
        f"max_rows_per_sheet: {max_rows}\n"
        "columns:\n"
        "  项目名称: A\n"
        "  数量: B\n"
        "  金额(USD): C\n"
    )
    path.write_text(payload, encoding="utf-8")
    return FixedMapping.from_yaml(path)


def test_write_fixed_splits_when_capacity_exceeded(tmp_path: Path) -> None:
    source = pd.DataFrame(
        [
            {"项目名称": "服务费", "数量": 1, "金额(USD)": 1200},
            {"项目名称": "备件", "数量": 2, "金额(USD)": 800},
            {"项目名称": "咨询", "数量": 3, "金额(USD)": 450},
        ]
    )
    source_path = tmp_path / "source.xlsx"
    source.to_excel(source_path, index=False)

    template_path = tmp_path / "template.xlsx"
    _build_template(template_path)

    mapping_path = tmp_path / "mapping.yaml"
    mapping = _write_mapping(mapping_path, max_rows=2)

    outputs = write_fixed(source, template_path, mapping, tmp_path / "invoice.xlsx")

    assert len(outputs) == 2
    first = outputs[0]
    second = outputs[1]
    assert first.exists() and second.exists()

    # Validate first workbook data placement
    from openpyxl import load_workbook

    wb1 = load_workbook(first)
    ws1 = wb1["Invoice"]
    assert ws1["A10"].value == "服务费"
    assert ws1["B11"].value == 2

    wb2 = load_workbook(second)
    ws2 = wb2["Invoice"]
    assert ws2["A10"].value == "咨询"


def test_mapping_raises_for_missing_columns(tmp_path: Path) -> None:
    df = pd.DataFrame([{"项目名称": "服务费", "数量": 1}])

    mapping_path = tmp_path / "mapping.yaml"
    mapping = _write_mapping(mapping_path, max_rows=5)
    target = TargetSchema(sheet="Invoice", start_row=10, header_row=9)

    with pytest.raises(MappingError):
        mapping.map(df, target)
