from __future__ import annotations

from pathlib import Path
from typing import Any
import datetime as dt

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook


def _ensure_template(template_path: Path) -> None:
    if template_path.exists():
        return
    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header rows
    ws["A1"] = "模板示例"
    ws["A2"] = "公司名称"
    ws["A3"] = "日期"
    ws["A4"] = "总金额"
    # Target cells for mapping
    ws["B2"] = ""
    ws["B3"] = ""
    ws["B4"] = 0
    wb.save(template_path)


def _load_mapping(mapping_path: Path) -> dict[str, Any]:
    if not mapping_path.exists():
        raise FileNotFoundError(f"映射文件不存在: {mapping_path}")
    with mapping_path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _eval_value(expr: Any, df: pd.DataFrame, profile: Any) -> Any:
    """Evaluate a simple mapping expression.

    Supported:
    - literal (number/string)
    - "today" -> yyyy-mm-dd
    - "sum:Column" -> sum of df[Column]
    - "first:Column" -> first non-null value
    - "$profile.xxx" -> attribute or key from profile
    """
    if expr is None:
        return None
    if isinstance(expr, (int, float)):
        return expr
    if not isinstance(expr, str):
        return str(expr)
    s = expr.strip()
    if s.lower() == "today":
        return dt.date.today().isoformat()
    if s.startswith("sum:"):
        col = s[4:]
        if col in df:
            return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())
        return 0
    if s.startswith("first:"):
        col = s[6:]
        if col in df:
            ser = df[col].dropna()
            return None if ser.empty else ser.iloc[0]
        return None
    if s.startswith("$profile."):
        key = s[len("$profile.") :]
        # dotted access
        parts = key.split(".")
        cur: Any = profile
        for p in parts:
            if hasattr(cur, p):
                cur = getattr(cur, p)
            elif isinstance(cur, dict) and p in cur:
                cur = cur[p]
            else:
                cur = None
                break
        return cur
    return s


def transform(
    input_path: Path,
    mapping_path: Path,
    template_path: Path,
    out_dir: Path,
    tmp_dir: Path,
    profile: Any,
) -> Path:
    """Read input xlsx, apply mapping rules, write into Excel template.

    Returns output file path.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    tmp_dir.mkdir(parents=True, exist_ok=True)

    _ensure_template(template_path)
    mapping = _load_mapping(mapping_path)

    df = pd.read_excel(input_path)

    # Optional simple cleaning per mapping
    ops = mapping.get("clean", {})
    if ops.get("dropna_columns"):
        df = df.dropna(subset=list(ops.get("dropna_columns", [])))
    if ops.get("fillna"):
        df = df.fillna(ops["fillna"])  # type: ignore[arg-type]

    wb = load_workbook(template_path)
    ws = wb.active

    cells: dict[str, Any] = mapping.get("cells", {})
    for cell, expr in cells.items():
        val = _eval_value(expr, df, profile)
        ws[cell] = val

    # Save intermediate for debugging
    tmp_output = tmp_dir / f"{profile.name}_intermediate.xlsx"
    wb.save(tmp_output)

    # Final output
    output = out_dir / f"{profile.name}_output.xlsx"
    wb.save(output)
    return output

