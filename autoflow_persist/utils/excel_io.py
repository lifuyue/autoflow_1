"""
RESPONSIBILITIES
- Manage atomic read/write operations for store workbooks via openpyxl.
- Provide lightweight locking to guard against concurrent writers.
PROCESS OVERVIEW
1. workbook_lock() acquires in-process and inter-process locks.
2. ensure_workbook() guarantees the sheet/header skeleton exists.
3. read_sheet() loads rows into dictionaries keyed by canonical columns.
4. write_sheet() writes rows back atomically using a temporary file swap.
"""

from __future__ import annotations

import os
import threading
from contextlib import contextmanager
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook

from autoflow_persist.stores.base_store import StoreLockedError

_IN_PROCESS_LOCKS: dict[Path, threading.RLock] = {}
_LOCK_REGISTRY_GUARD = threading.Lock()


def _acquire_inprocess_lock(path: Path) -> threading.RLock:
    with _LOCK_REGISTRY_GUARD:
        lock = _IN_PROCESS_LOCKS.get(path)
        if lock is None:
            lock = threading.RLock()
            _IN_PROCESS_LOCKS[path] = lock
        return lock


@contextmanager
def workbook_lock(path: Path) -> Iterable[None]:
    """Acquire a cooperative file lock guarding the given workbook."""

    path = path.resolve()
    inproc = _acquire_inprocess_lock(path)
    acquired = inproc.acquire(timeout=10)
    if not acquired:
        raise StoreLockedError(f"Timeout acquiring in-process lock for {path}")
    lock_path = path.with_suffix(path.suffix + ".lock")
    fd: int | None = None
    try:
        if lock_path.exists():
            raise StoreLockedError(f"Workbook appears locked: {lock_path}")
        fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(os.getpid()).encode("ascii"))
        yield
    except FileExistsError as exc:
        raise StoreLockedError(f"Workbook appears locked: {lock_path}") from exc
    finally:
        if fd is not None:
            os.close(fd)
        if lock_path.exists():
            os.unlink(lock_path)
        inproc.release()


def _tmp_path(path: Path) -> Path:
    return path.with_name(path.name + ".tmp")


def _atomic_save(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = _tmp_path(path)
    workbook.save(tmp_path)
    os.replace(tmp_path, path)


def ensure_workbook(path: Path, sheet_name: str, columns: Sequence[str]) -> None:
    """Ensure the target workbook and sheet with desired columns exists."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with workbook_lock(path):
        if not path.exists():
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            worksheet.append(list(columns))
            _atomic_save(workbook, path)
            return

        workbook = load_workbook(path)
        try:
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(title=sheet_name)
                worksheet.append(list(columns))
                _atomic_save(workbook, path)
                return

            worksheet = workbook[sheet_name]
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )
            current_header = [
                str(cell).strip() if cell is not None else ""
                for cell in (header_row or ())
            ]
            if current_header == list(columns):
                return

            header_map = {
                name: idx for idx, name in enumerate(current_header) if name
            }
            data_rows: list[dict[str, object]] = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                record = {
                    column: (
                        row[header_map[column]]
                        if column in header_map and header_map[column] < len(row)
                        else ""
                    )
                    for column in columns
                }
                data_rows.append(record)

            max_row = worksheet.max_row or 0
            if max_row:
                worksheet.delete_rows(1, max_row)
            worksheet.append(list(columns))
            for record in data_rows:
                worksheet.append([record.get(column, "") for column in columns])
            _atomic_save(workbook, path)
        finally:
            workbook.close()


def _read_without_lock(path: Path, sheet_name: str, columns: Sequence[str]) -> list[dict[str, object]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        header = [str(cell).strip() if cell is not None else "" for cell in header_row]
        index_map = {name: idx for idx, name in enumerate(header) if name}
        normalized_rows: list[dict[str, object]] = []
        for raw_values in rows_iter:
            if raw_values is None:
                continue
            if not any(cell is not None and str(cell).strip() for cell in raw_values):
                continue
            record: dict[str, object] = {}
            for column in columns:
                idx = index_map.get(column)
                if idx is None or idx >= len(raw_values):
                    record[column] = ""
                else:
                    value = raw_values[idx]
                    record[column] = "" if value is None else value
            normalized_rows.append(record)
        return normalized_rows
    finally:
        workbook.close()


def read_sheet(path: Path, sheet_name: str, columns: Sequence[str], *, use_lock: bool = True) -> list[dict[str, object]]:
    """Return worksheet content as dictionaries keyed by *columns*."""

    if use_lock:
        with workbook_lock(path):
            return _read_without_lock(path, sheet_name, columns)
    return _read_without_lock(path, sheet_name, columns)


def write_sheet(
    path: Path,
    sheet_name: str,
    rows: Iterable[Mapping[str, object]],
    columns: Sequence[str],
    *,
    use_lock: bool = True,
) -> None:
    """Write rows to a worksheet atomically."""

    def _write() -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet.append(list(columns))
        for row in rows:
            worksheet.append([row.get(column, "") for column in columns])
        _atomic_save(workbook, path)

    if use_lock:
        with workbook_lock(path):
            _write()
    else:
        _write()
