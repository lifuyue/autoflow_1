"""Excel output helpers for writing into templates."""

# Module responsibilities:
# - Apply mapping strategies to populate Excel templates without destroying styles.
# - Handle pagination when the template sheet reaches its configured capacity.

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .mapping import FixedMappingStrategy
from .schema import TargetSchema
from .utils.log import get_logger

logger = get_logger("excel_writer")


def _chunk_dataframe(df: pd.DataFrame, chunk_size: Optional[int]) -> Iterable[pd.DataFrame]:
    if not chunk_size or chunk_size <= 0:
        yield df
        return
    for start in range(0, len(df), chunk_size):
        yield df.iloc[start : start + chunk_size]


def _resolve_output_path(base: Path, index: int) -> Path:
    if index == 0:
        return base
    return base.with_name(f"{base.stem}_part{index + 1}{base.suffix}")


def _write_rows(
    ws: Worksheet,
    data: pd.DataFrame,
    column_map: dict[str, str],
    start_row: int,
) -> int:
    row_idx = start_row
    for _, row in data.iterrows():
        for source_col, target_col in column_map.items():
            cell_ref = f"{target_col}{row_idx}"
            ws[cell_ref].value = row[source_col]
        row_idx += 1
    return row_idx


def write_fixed(
    df: pd.DataFrame,
    template_path: Path,
    mapping: FixedMappingStrategy,
    out_path: Path,
    *,
    dry_run: bool = False,
) -> List[Path]:
    """Write a DataFrame into an Excel template using a fixed mapping strategy.

    Args:
        df: Source data validated elsewhere.
        template_path: Path to the template workbook.
        mapping: Fixed mapping strategy loaded from configuration.
        out_path: Output workbook path (first chunk); part files get ``_partN`` suffixes.
        dry_run: When True, skip actual file emission and only log the plan.

    Returns:
        List of generated output paths.

    Raises:
        FileNotFoundError: When the template workbook is absent.
        KeyError: When the target worksheet is missing.
    """

    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")

    target_schema = TargetSchema(
        sheet=mapping.sheet, start_row=mapping.start_row, header_row=mapping.header_row
    )
    column_map = mapping.map(df, target_schema)

    logger.info(
        "Starting Excel write",
        extra={
            "template": str(template_path),
            "rows": len(df),
            "columns": list(column_map.keys()),
            "max_rows_per_sheet": mapping.max_rows_per_sheet,
        },
    )

    outputs: List[Path] = []
    chunks = list(_chunk_dataframe(df, mapping.max_rows_per_sheet))
    for idx, chunk in enumerate(chunks):
        planned_out = _resolve_output_path(out_path, idx)
        outputs.append(planned_out)
        if dry_run:
            logger.info(
                "Dry run: would write chunk",
                extra={"rows": len(chunk), "output": str(planned_out)},
            )
            continue

        wb = load_workbook(template_path)
        if mapping.sheet not in wb.sheetnames:
            raise KeyError(f"Sheet '{mapping.sheet}' not found in template")
        ws = wb[mapping.sheet]
        _write_rows(ws, chunk, column_map, mapping.start_row)
        planned_out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(planned_out)
        logger.info(
            "Chunk written",
            extra={"rows": len(chunk), "output": str(planned_out)},
        )

    if len(outputs) > 1:
        logger.info(
            "Data split across multiple output files due to row capacity",
            extra={"files": [str(p) for p in outputs]},
        )

    return outputs
